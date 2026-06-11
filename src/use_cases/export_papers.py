import csv
import os
import tempfile
import zipfile
from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter, range_boundaries

from src.domain.criteria_config import DEFAULT_CRITERIA
from src.domain.enums import CriterionType, ScreeningStatus
from src.domain.interfaces import PaperRepository, ScreeningDecisionRepository

_HEADER_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
_HEADER_FONT = Font(name="Segoe UI", size=11, bold=True, color="333333")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
_DATA_FONT = Font(name="Segoe UI", size=10, color="333333")
_THIN_BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)
_CENTER_ALIGN = Alignment(horizontal="center", vertical="top")
_WRAP_ALIGN = Alignment(wrap_text=True, vertical="top", horizontal="left")

_GREEN_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
_GREEN_FONT = Font(name="Segoe UI", size=10, color="375623", bold=True)
_ORANGE_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
_ORANGE_FONT = Font(name="Segoe UI", size=10, color="C65911", bold=True)
_YELLOW_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
_YELLOW_FONT = Font(name="Segoe UI", size=10, color="806000", bold=True)


def _apply_status_style(cell, value):
    if value in ("YES", 1):
        cell.fill = _GREEN_FILL
        cell.font = _GREEN_FONT
    elif value in ("NO", 0):
        cell.fill = _ORANGE_FILL
        cell.font = _ORANGE_FONT
    else:
        cell.fill = _YELLOW_FILL
        cell.font = _YELLOW_FONT


def style_range(ws, cell_range, border=None, fill=None, alignment=None):
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            if border:
                cell.border = border
            if fill:
                cell.fill = fill
            if alignment:
                cell.alignment = alignment

WL_HEADERS = [
    "Library",
    "",
    "#",
    "Title",
    "Abstract",
    "Palavra-Chaves",
    "CIs1",
    "CEs1",
    "Revisor 1",
    "CIs2",
    "CEs2",
    "Revisor 2",
    "Decision",
]

GL_HEADERS = [
    "Library",
    "",
    "#",
    "Title",
    "Abstract",
    "Palavra-Chaves",
    "CIs1",
    "CEs1",
    "Revisor 1",
    "CIs2",
    "CEs2",
    "Revisor 2",
    "Decision",
]


def _get_inclusion_criteria_text() -> str:
    criteria = [c for c in DEFAULT_CRITERIA if c.type == CriterionType.INCLUSION]
    return "\n".join(f"{c.code}: {c.description}" for c in criteria)


def _get_exclusion_criteria_text() -> str:
    criteria = [c for c in DEFAULT_CRITERIA if c.type == CriterionType.EXCLUSION]
    return "\n".join(f"{c.code}: {c.description}" for c in criteria)


def _get_full_protocol_text() -> str:
    ic_text = _get_inclusion_criteria_text()
    ec_text = _get_exclusion_criteria_text()
    return f"Inclusion Criteria (IC):\n{ic_text}\n\nExclusion Criteria (EC):\n{ec_text}"


def _get_library(paper) -> str:
    return paper.metadata.get("Detected_Source") or paper.metadata.get("Library") or "Unknown Database"


def _get_keywords(paper) -> str:
    keywords = paper.metadata.get("Keywords") or paper.metadata.get("Palavra-Chaves")
    if isinstance(keywords, list):
        return "; ".join(keywords)
    return keywords or ""


def _map_status_to_revisor(status: ScreeningStatus) -> str:
    if status == ScreeningStatus.INCLUDED:
        return "YES"
    elif status == ScreeningStatus.EXCLUDED:
        return "NO"
    return "NEEDS_REVIEW"


def _map_status_to_numeric(status: ScreeningStatus) -> int | str:
    if status == ScreeningStatus.INCLUDED:
        return 1
    elif status == ScreeningStatus.EXCLUDED:
        return 0
    return "NEEDS_REVIEW"


class ExportScreenedPapersUseCase:
    def __init__(
        self,
        paper_repository: PaperRepository,
        decision_repository: ScreeningDecisionRepository,
    ) -> None:
        self._paper_repo = paper_repository
        self._decision_repo = decision_repository

    def execute(self, output_path: str = "data/processed/apollo_results.xlsx") -> None:
        papers = self._paper_repo.get_all_papers()
        decisions = self._decision_repo.get_all_decisions()
        decisions = [d for d in decisions if d.status in (ScreeningStatus.INCLUDED, ScreeningStatus.EXCLUDED)]
        paper_map = {p.id: p for p in papers}
        human_map = self._decision_repo.get_human_decision_map()

        wl_papers = []
        gl_papers = []
        for decision in decisions:
            paper = paper_map.get(decision.paper_id)
            if paper:
                if paper.source_type.value == "WL":
                    wl_papers.append((paper, decision))
                else:
                    gl_papers.append((paper, decision))

        out = Path(output_path)
        os.makedirs(str(out.parent), exist_ok=True)

        wb = Workbook()
        wb.remove(wb.active)

        quality_map = self._decision_repo.get_quality_map()

        self._create_wl_sheet(wb, wl_papers, human_map)
        self._create_gl_sheet(wb, gl_papers, human_map, quality_map)

        wb.save(str(out))

    def _create_wl_sheet(self, wb, wl_papers: list, human_map: dict[str, str]) -> None:
        ws = wb.create_sheet("White Literature")
        ws.views.sheetView[0].showGridLines = True
        ws.row_dimensions[1].height = 28

        for col_idx, header in enumerate(WL_HEADERS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = _HEADER_ALIGN
            cell.border = _THIN_BORDER

        ic_comment_text = _get_inclusion_criteria_text()
        ec_comment_text = _get_exclusion_criteria_text()

        for cell_ref in ("G1", "J1"):
            c = Comment(ic_comment_text, "APOLLO")
            c.width = 350
            c.height = 220
            ws[cell_ref].comment = c
        for cell_ref in ("H1", "K1"):
            c = Comment(ec_comment_text, "APOLLO")
            c.width = 350
            c.height = 220
            ws[cell_ref].comment = c

        for row_idx, (paper, decision) in enumerate(wl_papers, start=2):
            ws.row_dimensions[row_idx].height = 20
            ws.cell(row=row_idx, column=1, value=_get_library(paper))
            ws.cell(row=row_idx, column=2, value=paper.id)
            ws.cell(row=row_idx, column=3, value=row_idx - 1)
            ws.cell(row=row_idx, column=4, value=paper.title)
            ws.cell(row=row_idx, column=5, value=paper.abstract or "")
            ws.cell(row=row_idx, column=6, value=_get_keywords(paper))

            inclusion_codes = [c for c in decision.applied_criteria_codes if c.startswith("IC")]
            exclusion_codes = [c for c in decision.applied_criteria_codes if c.startswith("EC")]

            ws.cell(row=row_idx, column=7, value="; ".join(inclusion_codes) if inclusion_codes else "")
            ws.cell(row=row_idx, column=8, value="; ".join(exclusion_codes) if exclusion_codes else "")

            human_val = human_map.get(paper.id)
            if human_val:
                revisor_val = human_val
                _apply_status_style(ws.cell(row=row_idx, column=9), human_val)
            else:
                revisor_val = ""
            ws.cell(row=row_idx, column=9, value=revisor_val)

            for col_idx in range(1, 14):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = _DATA_FONT
                cell.border = _THIN_BORDER
                if col_idx in (1, 2, 3, 7, 8, 9, 10, 11, 12, 13):
                    cell.alignment = _CENTER_ALIGN
                else:
                    cell.alignment = _WRAP_ALIGN

        wl_col_widths = {1: 18, 2: 12, 3: 6, 4: 50, 5: 75, 6: 25,
                         7: 10, 8: 10, 9: 15, 10: 10, 11: 10, 12: 15, 13: 15}
        for col, width in wl_col_widths.items():
            ws.column_dimensions[get_column_letter(col)].width = width

    def _create_gl_sheet(self, wb, gl_papers: list, human_map: dict[str, str], quality_map: dict[str, dict]) -> None:
        ws = wb.create_sheet("Grey Literature")
        ws.views.sheetView[0].showGridLines = True
        ws.row_dimensions[1].height = 28

        for col_idx, header in enumerate(GL_HEADERS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = _HEADER_ALIGN
            cell.border = _THIN_BORDER

        ic_comment_text = _get_inclusion_criteria_text()
        ec_comment_text = _get_exclusion_criteria_text()

        for cell_ref in ("G1", "J1"):
            c = Comment(ic_comment_text, "APOLLO")
            c.width = 350
            c.height = 220
            ws[cell_ref].comment = c
        for cell_ref in ("H1", "K1"):
            c = Comment(ec_comment_text, "APOLLO")
            c.width = 350
            c.height = 220
            ws[cell_ref].comment = c

        for row_idx, (paper, decision) in enumerate(gl_papers, start=2):
            ws.row_dimensions[row_idx].height = 20
            ws.cell(row=row_idx, column=1, value=_get_library(paper))
            ws.cell(row=row_idx, column=2, value=paper.id)
            ws.cell(row=row_idx, column=3, value=row_idx - 1)
            ws.cell(row=row_idx, column=4, value=paper.title)

            crawled = quality_map.get(paper.id, {}).get("full_text")
            ws.cell(row=row_idx, column=5, value=crawled or "(No web content scraped)")

            ws.cell(row=row_idx, column=6, value=_get_keywords(paper))

            inclusion_codes = [c for c in decision.applied_criteria_codes if c.startswith("IC")]
            exclusion_codes = [c for c in decision.applied_criteria_codes if c.startswith("EC")]

            ws.cell(row=row_idx, column=7, value="; ".join(inclusion_codes) if inclusion_codes else "")
            ws.cell(row=row_idx, column=8, value="; ".join(exclusion_codes) if exclusion_codes else "")

            human_val = human_map.get(paper.id)
            if human_val:
                revisor_val = human_val
                _apply_status_style(ws.cell(row=row_idx, column=9), human_val)
            else:
                revisor_val = ""
            ws.cell(row=row_idx, column=9, value=revisor_val)

            for col_idx in range(1, 14):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = _DATA_FONT
                cell.border = _THIN_BORDER
                if col_idx in (1, 2, 3, 7, 8, 9, 10, 11, 12, 13):
                    cell.alignment = _CENTER_ALIGN
                else:
                    cell.alignment = _WRAP_ALIGN

        gl_col_widths = {1: 18, 2: 12, 3: 6, 4: 50, 5: 75, 6: 25,
                         7: 10, 8: 10, 9: 15, 10: 10, 11: 10, 12: 15, 13: 15}
        for col, width in gl_col_widths.items():
            ws.column_dimensions[get_column_letter(col)].width = width

    def export_as_csv_zip(self) -> Path:
        papers = self._paper_repo.get_all_papers()
        decisions = self._decision_repo.get_all_decisions()
        decisions = [d for d in decisions if d.status in (ScreeningStatus.INCLUDED, ScreeningStatus.EXCLUDED)]
        paper_map = {p.id: p for p in papers}
        human_map = self._decision_repo.get_human_decision_map()
        quality_map = self._decision_repo.get_quality_map()

        wl_rows = []
        gl_rows = []
        for decision in decisions:
            paper = paper_map.get(decision.paper_id)
            if paper:
                row = self._build_csv_row(paper, decision, human_map, quality_map)
                if paper.source_type.value == "WL":
                    wl_rows.append(row)
                else:
                    gl_rows.append(row)

        tmp_dir = Path(tempfile.mkdtemp())
        wl_path = tmp_dir / "White_Literature_Results.csv"
        gl_path = tmp_dir / "Grey_Literature_Results.csv"
        zip_path = tmp_dir / "APOLLO_Screening_Results.zip"

        HEADERS = [
            "Library", "", "#", "Title", "Abstract", "Palavra-Chaves",
            "CIs1", "CEs1", "Revisor 1", "CIs2", "CEs2", "Revisor 2", "Decision",
        ]

        with open(wl_path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(HEADERS)
            for idx, row in enumerate(wl_rows, start=1):
                row[2] = str(idx)
                writer.writerow(row)

        with open(gl_path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(HEADERS)
            for idx, row in enumerate(gl_rows, start=1):
                row[2] = str(idx)
                writer.writerow(row)

        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
            zf.write(wl_path, "White_Literature_Results.csv")
            zf.write(gl_path, "Grey_Literature_Results.csv")

        wl_path.unlink()
        gl_path.unlink()
        return zip_path

    def _build_csv_row(self, paper, decision, human_map, quality_map) -> list[str]:
        inclusion_codes = [c for c in decision.applied_criteria_codes if c.startswith("IC")]
        exclusion_codes = [c for c in decision.applied_criteria_codes if c.startswith("EC")]

        if paper.source_type.value == "GL":
            crawled = quality_map.get(paper.id, {}).get("full_text")
            abstract = crawled or "(No web content scraped)"
        else:
            abstract = paper.abstract or ""

        human_val = human_map.get(paper.id, "")

        final_decision = _map_status_to_revisor(decision.status)

        return [
            _get_library(paper),
            paper.id,
            str(0),
            paper.title,
            abstract,
            _get_keywords(paper),
            "; ".join(inclusion_codes) if inclusion_codes else "",
            "; ".join(exclusion_codes) if exclusion_codes else "",
            human_val,
            "",
            "",
            "",
            final_decision,
        ]
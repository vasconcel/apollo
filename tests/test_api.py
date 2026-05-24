import os
from pathlib import Path
from unittest.mock import AsyncMock, MagicMock, patch

import pytest
from fastapi.testclient import TestClient

from src.api.main import app
from src.domain.enums import ScreeningStatus, SourceType
from src.domain.models import Paper, ScreeningDecision

client = TestClient(app)


@pytest.fixture(autouse=True)
def _reset_global_state():
    from src.api import routes
    routes._dataset_path = None
    routes._screening_active = False
    routes._imported_count = 0
    yield


@pytest.fixture
def test_env(tmp_path, monkeypatch):
    from src.api import routes

    import_dir = tmp_path / "imported"
    process_dir = tmp_path / "processed"
    import_dir.mkdir()
    process_dir.mkdir()
    db_path = str(tmp_path / "screening.db")

    monkeypatch.setattr(routes, "_IMPORT_DIR", import_dir)
    monkeypatch.setattr(routes, "_PROCESSED_DIR", process_dir)
    monkeypatch.setenv("APOLLO_DB_PATH", db_path)

    return tmp_path


CSV_CONTENT = (
    "Global_ID,Title,Abstract,Provenance_Trace\n"
    "1,Paper One,Abstract One,wl:acm\n"
    "2,Paper Two,Abstract Two,gl:google\n"
    "3,Paper Three,Abstract Three,wl:ieee\n"
)

XLSX_ROW_COUNT = 3


# ── POST /api/import ────────────────────────────────────────────────────────


class TestImport:
    def test_import_csv_success(self, test_env):
        resp = client.post(
            "/api/import",
            files={"file": ("test.csv", CSV_CONTENT, "text/csv")},
        )
        assert resp.status_code == 200
        data = resp.json()
        assert data["status"] == "success"
        assert data["imported_count"] == XLSX_ROW_COUNT
        assert data["skipped_duplicates"] == 0

    def test_import_no_file_returns_422(self, test_env):
        resp = client.post("/api/import")
        assert resp.status_code == 422

    def test_import_twice_updates_count(self, test_env):
        csv2 = "Global_ID,Title,Abstract,Provenance_Trace\n9,Extra Paper,Extra Abstract,wl:test\n"
        resp = client.post(
            "/api/import",
            files={"file": ("a.csv", CSV_CONTENT, "text/csv")},
        )
        assert resp.json()["imported_count"] == XLSX_ROW_COUNT

        resp2 = client.post(
            "/api/import",
            files={"file": ("b.csv", csv2, "text/csv")},
        )
        assert resp2.json()["imported_count"] == 1


# ── POST /api/screening/start ───────────────────────────────────────────────


class TestScreeningStart:
    def test_start_without_import_returns_400(self, test_env):
        resp = client.post("/api/screening/start")
        assert resp.status_code == 400
        assert "No dataset imported" in resp.json()["detail"]

    def test_start_returns_started(self, test_env):
        client.post(
            "/api/import",
            files={"file": ("test.csv", CSV_CONTENT, "text/csv")},
        )

        with patch(
            "src.api.routes._run_screening_background",
            new=AsyncMock(),
        ):
            resp = client.post("/api/screening/start")
        assert resp.status_code == 200
        data = resp.json()
        assert data["status"] == "started"
        assert data["task_id"] == "screening_current"

    def test_start_while_active_returns_409(self, test_env):
        from src.api import routes
        client.post(
            "/api/import",
            files={"file": ("test.csv", CSV_CONTENT, "text/csv")},
        )
        routes._screening_active = True
        try:
            resp = client.post("/api/screening/start")
            assert resp.status_code == 409
        finally:
            routes._screening_active = False


# ── GET /api/screening/progress ─────────────────────────────────────────────


class TestScreeningProgress:
    def test_progress_without_import_returns_zeros(self, test_env):
        resp = client.get("/api/screening/progress")
        assert resp.status_code == 200
        data = resp.json()
        assert data["total_papers"] == 0
        assert data["screened_count"] == 0
        assert data["is_active"] is False

    def test_progress_after_import_without_screening(self, test_env):
        client.post(
            "/api/import",
            files={"file": ("test.csv", CSV_CONTENT, "text/csv")},
        )
        resp = client.get("/api/screening/progress")
        assert resp.status_code == 200
        data = resp.json()
        assert data["total_papers"] == XLSX_ROW_COUNT
        assert data["screened_count"] == 0
        assert data["pending_count"] == XLSX_ROW_COUNT

    def test_progress_counts_decisions(self, test_env):
        client.post(
            "/api/import",
            files={"file": ("test.csv", CSV_CONTENT, "text/csv")},
        )

        from src.api.routes import _get_decision_repo
        repo = _get_decision_repo()
        repo.save_decision(ScreeningDecision(
            paper_id="1",
            status=ScreeningStatus.EXCLUDED,
            confidence_score=1.0,
            rationale="Excluded by EC4",
            applied_criteria_codes=["EC4"],
        ))
        repo.save_decision(ScreeningDecision(
            paper_id="2",
            status=ScreeningStatus.EXCLUDED,
            confidence_score=1.0,
            rationale="Excluded by LLM",
            applied_criteria_codes=["EC2"],
        ))

        resp = client.get("/api/screening/progress")
        assert resp.status_code == 200
        data = resp.json()
        assert data["total_papers"] == XLSX_ROW_COUNT
        assert data["screened_count"] == 2
        assert data["pending_count"] == 1
        assert data["heuristic_exclusions"] == 1
        assert data["ai_exclusions"] == 1


# ── GET /api/papers ─────────────────────────────────────────────────────────


class TestPapers:
    def _import_and_add_decision(self, decision: ScreeningDecision):
        client.post(
            "/api/import",
            files={"file": ("test.csv", CSV_CONTENT, "text/csv")},
        )
        from src.api.routes import _get_decision_repo
        _get_decision_repo().save_decision(decision)

    def test_papers_returns_paginated_results(self, test_env):
        self._import_and_add_decision(ScreeningDecision(
            paper_id="1",
            status=ScreeningStatus.INCLUDED,
            confidence_score=0.95,
            rationale="Good paper",
            applied_criteria_codes=["IC1", "IC2"],
        ))
        resp = client.get("/api/papers?page=1&size=2")
        assert resp.status_code == 200
        data = resp.json()
        assert data["page"] == 1
        assert data["size"] == 2
        assert data["total"] == XLSX_ROW_COUNT
        assert len(data["items"]) == 2

    def test_papers_second_page(self, test_env):
        self._import_and_add_decision(ScreeningDecision(
            paper_id="1",
            status=ScreeningStatus.INCLUDED,
            confidence_score=0.95,
            rationale="Good paper",
            applied_criteria_codes=["IC1"],
        ))
        resp = client.get("/api/papers?page=2&size=2")
        assert resp.status_code == 200
        data = resp.json()
        assert len(data["items"]) == XLSX_ROW_COUNT - 2

    def test_papers_filter_by_status(self, test_env):
        self._import_and_add_decision(ScreeningDecision(
            paper_id="1",
            status=ScreeningStatus.INCLUDED,
            confidence_score=0.95,
            rationale="Good",
            applied_criteria_codes=["IC1"],
        ))
        resp = client.get("/api/papers?status=INCLUDED")
        assert resp.status_code == 200
        data = resp.json()
        assert data["total"] == 1
        assert data["items"][0]["status"] == "INCLUDED"

    def test_papers_filter_no_match(self, test_env):
        self._import_and_add_decision(ScreeningDecision(
            paper_id="1",
            status=ScreeningStatus.INCLUDED,
            confidence_score=0.95,
            rationale="Good",
            applied_criteria_codes=["IC1"],
        ))
        resp = client.get("/api/papers?status=NEEDS_REVIEW")
        assert resp.status_code == 200
        data = resp.json()
        assert data["total"] == 0

    def test_papers_without_import_returns_400(self, test_env):
        resp = client.get("/api/papers")
        assert resp.status_code == 400

    def test_paper_fields_structure(self, test_env):
        self._import_and_add_decision(ScreeningDecision(
            paper_id="1",
            status=ScreeningStatus.INCLUDED,
            confidence_score=0.95,
            rationale="Good paper",
            applied_criteria_codes=["IC1", "IC2"],
        ))
        resp = client.get("/api/papers?page=1&size=1")
        assert resp.status_code == 200
        item = resp.json()["items"][0]
        assert "id" in item
        assert "title" in item
        assert "abstract" in item
        assert "source_library" in item
        assert "source_type" in item
        assert "publication_year" in item
        assert "url" in item
        assert "status" in item
        assert "rationale" in item
        assert "confidence_score" in item
        assert "applied_criteria_codes" in item


# ── GET /api/export ─────────────────────────────────────────────────────────


class TestExport:
    def test_export_without_import_returns_400(self, test_env):
        resp = client.get("/api/export")
        assert resp.status_code == 400

    def test_export_returns_xlsx(self, test_env):
        client.post(
            "/api/import",
            files={"file": ("test.csv", CSV_CONTENT, "text/csv")},
        )
        from src.api.routes import _get_decision_repo
        repo = _get_decision_repo()
        repo.save_decision(ScreeningDecision(
            paper_id="1",
            status=ScreeningStatus.INCLUDED,
            confidence_score=0.95,
            rationale="Good paper",
            applied_criteria_codes=["IC1"],
        ))
        repo.save_decision(ScreeningDecision(
            paper_id="2",
            status=ScreeningStatus.EXCLUDED,
            confidence_score=1.0,
            rationale="Bad paper",
            applied_criteria_codes=["EC6"],
        ))

        resp = client.get("/api/export")
        assert resp.status_code == 200
        ct = resp.headers["content-type"]
        assert "spreadsheetml.sheet" in ct
        assert resp.headers["content-disposition"] == 'attachment; filename="APOLLO_Screening_Results.xlsx"'

        from openpyxl import load_workbook
        wb = load_workbook(Path(test_env) / "processed" / "export" / "APOLLO_Screening_Results.xlsx")
        assert "White Literature" in wb.sheetnames
        assert "Grey Literature" in wb.sheetnames

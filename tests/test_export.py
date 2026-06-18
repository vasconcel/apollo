import pytest
from openpyxl import load_workbook
from pytest_mock import MockerFixture

from src.domain.enums import ScreeningStatus, SourceType
from src.domain.models import Paper, ScreeningDecision


@pytest.fixture
def mock_paper_repo(mocker: MockerFixture):
    repo = mocker.MagicMock()
    repo.get_all_papers.return_value = [
        Paper(
            id="1",
            title="Paper One",
            source_type=SourceType.WL,
            publication_year=2023,
            url="https://example.com/1",
            abstract="Abstract one.",
            metadata={"Detected_Source": "ACM Digital Library", "Keywords": "agile; review"},
        ),
        Paper(
            id="2",
            title="Paper Two",
            source_type=SourceType.GL,
            publication_year=2022,
            url="https://example.com/2",
            abstract="Abstract two.",
            metadata={},
        ),
    ]
    return repo


@pytest.fixture
def mock_decision_repo(mocker: MockerFixture):
    repo = mocker.MagicMock()
    repo.get_all_decisions.return_value = [
        ScreeningDecision(
            paper_id="1",
            status=ScreeningStatus.INCLUDED,
            confidence_score=0.95,
            rationale="**EC Analysis:** ...\n**IC Analysis:** ...\n**Conclusion:** Good paper.",
            applied_criteria_codes=["IC1", "IC2"],
        ),
        ScreeningDecision(
            paper_id="2",
            status=ScreeningStatus.EXCLUDED,
            confidence_score=1.0,
            rationale="Duplicate detected.",
            applied_criteria_codes=["EC6"],
        ),
    ]
    repo.get_human_decision_map.return_value = {}
    repo.get_quality_map.return_value = {}
    return repo


class TestFileAndSheets:
    def test_both_sheets_exist(self, mock_paper_repo, mock_decision_repo, tmp_path):
        from src.use_cases.export_papers import ExportScreenedPapersUseCase

        output = tmp_path / "output.xlsx"
        ExportScreenedPapersUseCase(mock_paper_repo, mock_decision_repo).execute(str(output))

        wb = load_workbook(output)
        assert "White Literature" in wb.sheetnames
        assert "Grey Literature" in wb.sheetnames

    def test_no_ai_columns(self, mock_paper_repo, mock_decision_repo, tmp_path):
        from src.use_cases.export_papers import ExportScreenedPapersUseCase

        output = tmp_path / "output.xlsx"
        ExportScreenedPapersUseCase(mock_paper_repo, mock_decision_repo).execute(str(output))

        wb = load_workbook(output)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            headers = [str(c.value or "") for c in ws[1]]
            assert "AI_Decision" not in headers
            assert "AI_Rationale" not in headers


class TestWLSheet:
    WL_EXPECTED_HEADERS = [
        "Library", "", "#", "Title", "Abstract",
        "Palavra-Chaves", "CIs1", "CEs1", "Revisor 1",
        "CIs2", "CEs2", "Revisor 2", "Decision",
    ]

    def test_headers_exact_order(self, mock_paper_repo, mock_decision_repo, tmp_path):
        from src.use_cases.export_papers import ExportScreenedPapersUseCase

        output = tmp_path / "output.xlsx"
        ExportScreenedPapersUseCase(mock_paper_repo, mock_decision_repo).execute(str(output))

        wb = load_workbook(output)
        ws = wb["White Literature"]
        headers = [c.value for c in ws[1]]
        assert headers[0] == "Library"
        assert headers[1] is None
        assert headers[2:] == self.WL_EXPECTED_HEADERS[2:]

    def test_row_content(self, mock_paper_repo, mock_decision_repo, tmp_path):
        from src.use_cases.export_papers import ExportScreenedPapersUseCase

        output = tmp_path / "output.xlsx"
        ExportScreenedPapersUseCase(mock_paper_repo, mock_decision_repo).execute(str(output))

        wb = load_workbook(output)
        ws = wb["White Literature"]

        assert ws.cell(row=2, column=1).value == "ACM Digital Library"
        assert ws.cell(row=2, column=2).value == "1"
        assert ws.cell(row=2, column=3).value == 1
        assert ws.cell(row=2, column=4).value == "Paper One"
        assert ws.cell(row=2, column=5).value == "Abstract one."
        assert ws.cell(row=2, column=6).value == "agile; review"
        assert ws.cell(row=2, column=7).value == "IC1; IC2"
        assert ws.cell(row=2, column=8).value is None
        assert ws.cell(row=2, column=9).value is None

    def test_human_columns_empty_in_data(self, mock_paper_repo, mock_decision_repo, tmp_path):
        from src.use_cases.export_papers import ExportScreenedPapersUseCase

        output = tmp_path / "output.xlsx"
        ExportScreenedPapersUseCase(mock_paper_repo, mock_decision_repo).execute(str(output))

        wb = load_workbook(output)
        ws = wb["White Literature"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            assert row[9] is None or row[9] == ""
            assert row[10] is None or row[10] == ""
            assert row[11] is None or row[11] == ""
            assert row[12] is None or row[12] == ""

    def test_cis1_comment_exists(self, mock_paper_repo, mock_decision_repo, tmp_path):
        from src.use_cases.export_papers import ExportScreenedPapersUseCase

        output = tmp_path / "output.xlsx"
        ExportScreenedPapersUseCase(mock_paper_repo, mock_decision_repo).execute(str(output))

        wb = load_workbook(output)
        ws = wb["White Literature"]
        comment = ws["G1"].comment
        assert comment is not None
        assert "IC1" in comment.text
        assert "IC2" in comment.text
        assert "IC3" in comment.text

    def test_cis2_comment_exists(self, mock_paper_repo, mock_decision_repo, tmp_path):
        from src.use_cases.export_papers import ExportScreenedPapersUseCase

        output = tmp_path / "output.xlsx"
        ExportScreenedPapersUseCase(mock_paper_repo, mock_decision_repo).execute(str(output))

        wb = load_workbook(output)
        ws = wb["White Literature"]
        comment = ws["J1"].comment
        assert comment is not None
        assert "IC1" in comment.text

    def test_ces1_comment_exists(self, mock_paper_repo, mock_decision_repo, tmp_path):
        from src.use_cases.export_papers import ExportScreenedPapersUseCase

        output = tmp_path / "output.xlsx"
        ExportScreenedPapersUseCase(mock_paper_repo, mock_decision_repo).execute(str(output))

        wb = load_workbook(output)
        ws = wb["White Literature"]
        comment = ws["H1"].comment
        assert comment is not None
        assert "EC1" in comment.text
        assert "EC6" in comment.text

    def test_ces2_comment_exists(self, mock_paper_repo, mock_decision_repo, tmp_path):
        from src.use_cases.export_papers import ExportScreenedPapersUseCase

        output = tmp_path / "output.xlsx"
        ExportScreenedPapersUseCase(mock_paper_repo, mock_decision_repo).execute(str(output))

        wb = load_workbook(output)
        ws = wb["White Literature"]
        comment = ws["K1"].comment
        assert comment is not None
        assert "EC1" in comment.text

    def test_missing_wl_metadata_defaults(self, mocker: MockerFixture, tmp_path):
        from src.use_cases.export_papers import ExportScreenedPapersUseCase

        repo = mocker.MagicMock()
        repo.get_all_papers.return_value = [
            Paper(
                id="10",
                title="No Metadata",
                source_type=SourceType.WL,
                publication_year=2023,
                metadata={},
            ),
        ]
        repo.get_all_decisions.return_value = [
            ScreeningDecision(
                paper_id="10",
                status=ScreeningStatus.INCLUDED,
                confidence_score=0.95,
                rationale="Meets criteria",
                applied_criteria_codes=["IC1"],
            ),
        ]
        repo.get_human_decision_map.return_value = {}

        output = tmp_path / "output.xlsx"
        ExportScreenedPapersUseCase(repo, repo).execute(str(output))

        wb = load_workbook(output)
        ws = wb["White Literature"]
        headers = [c.value for c in ws[1]]
        row = {headers[i]: ws.cell(row=2, column=i + 1).value for i in range(len(headers))}

        assert row["Library"] == "Unknown Database"
        assert row["Palavra-Chaves"] is None
        assert row["Revisor 1"] is None


class TestGLSheet:
    GL_EXPECTED_HEADERS = [
        "Library", "", "#", "Title", "Web Content",
        "Palavra-Chaves", "CIs1", "CEs1", "Revisor 1",
        "CIs2", "CEs2", "Revisor 2", "Decision",
    ]

    def test_headers_exact_order(self, mock_paper_repo, mock_decision_repo, tmp_path):
        from src.use_cases.export_papers import ExportScreenedPapersUseCase

        output = tmp_path / "output.xlsx"
        ExportScreenedPapersUseCase(mock_paper_repo, mock_decision_repo).execute(str(output))

        wb = load_workbook(output)
        ws = wb["Grey Literature"]
        headers = [c.value for c in ws[1]]
        assert headers[0] == "Library"
        assert headers[1] is None
        assert headers[2:] == self.GL_EXPECTED_HEADERS[2:]

    def test_no_merged_cells(self, mock_paper_repo, mock_decision_repo, tmp_path):
        from src.use_cases.export_papers import ExportScreenedPapersUseCase

        output = tmp_path / "output.xlsx"
        ExportScreenedPapersUseCase(mock_paper_repo, mock_decision_repo).execute(str(output))

        wb = load_workbook(output)
        ws = wb["Grey Literature"]
        assert len(ws.merged_cells.ranges) == 0

    def test_row_content(self, mock_paper_repo, mock_decision_repo, tmp_path):
        from src.use_cases.export_papers import ExportScreenedPapersUseCase

        output = tmp_path / "output.xlsx"
        ExportScreenedPapersUseCase(mock_paper_repo, mock_decision_repo).execute(str(output))

        wb = load_workbook(output)
        ws = wb["Grey Literature"]

        assert ws.cell(row=2, column=1).value == "Unknown Database"
        assert ws.cell(row=2, column=2).value == "2"
        assert ws.cell(row=2, column=3).value == 1
        assert ws.cell(row=2, column=4).value == "Paper Two"
        assert ws.cell(row=2, column=5).value == "(No web content scraped)"
        assert ws.cell(row=2, column=6).value in ("", None)
        assert ws.cell(row=2, column=7).value in ("", None)
        assert ws.cell(row=2, column=8).value == "EC6"
        assert ws.cell(row=2, column=9).value in ("", None)

    def test_gl_abstract_shows_crawled_content(self, mocker: MockerFixture, tmp_path):
        from src.use_cases.export_papers import ExportScreenedPapersUseCase

        repo = mocker.MagicMock()
        repo.get_all_papers.return_value = [
            Paper(
                id="gl-1",
                title="GL Paper With Crawl",
                source_type=SourceType.GL,
                publication_year=2023,
                abstract="Original abstract (should not appear)",
            ),
        ]
        repo.get_all_decisions.return_value = [
            ScreeningDecision(
                paper_id="gl-1",
                status=ScreeningStatus.INCLUDED,
                confidence_score=0.9,
                rationale="Good",
                applied_criteria_codes=["IC1"],
            ),
        ]
        repo.get_human_decision_map.return_value = {}
        repo.get_quality_map.return_value = {
            "gl-1": {"full_text": "Scraped web content for GL paper"},
        }

        output = tmp_path / "output.xlsx"
        ExportScreenedPapersUseCase(repo, repo).execute(str(output))

        wb = load_workbook(output)
        ws = wb["Grey Literature"]
        assert ws.cell(row=2, column=5).value == "Scraped web content for GL paper"

    def test_human_columns_empty_in_data(self, mock_paper_repo, mock_decision_repo, tmp_path):
        from src.use_cases.export_papers import ExportScreenedPapersUseCase

        output = tmp_path / "output.xlsx"
        ExportScreenedPapersUseCase(mock_paper_repo, mock_decision_repo).execute(str(output))

        wb = load_workbook(output)
        ws = wb["Grey Literature"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            assert row[9] is None or row[9] == ""
            assert row[10] is None or row[10] == ""
            assert row[11] is None or row[11] == ""
            assert row[12] is None or row[12] == ""

    def test_cis1_comment_exists(self, mock_paper_repo, mock_decision_repo, tmp_path):
        from src.use_cases.export_papers import ExportScreenedPapersUseCase

        output = tmp_path / "output.xlsx"
        ExportScreenedPapersUseCase(mock_paper_repo, mock_decision_repo).execute(str(output))

        wb = load_workbook(output)
        ws = wb["Grey Literature"]
        comment = ws["G1"].comment
        assert comment is not None
        assert "IC1" in comment.text
        assert "IC2" in comment.text
        assert "IC3" in comment.text

    def test_cis2_comment_exists(self, mock_paper_repo, mock_decision_repo, tmp_path):
        from src.use_cases.export_papers import ExportScreenedPapersUseCase

        output = tmp_path / "output.xlsx"
        ExportScreenedPapersUseCase(mock_paper_repo, mock_decision_repo).execute(str(output))

        wb = load_workbook(output)
        ws = wb["Grey Literature"]
        comment = ws["J1"].comment
        assert comment is not None
        assert "IC1" in comment.text

    def test_ces1_comment_exists(self, mock_paper_repo, mock_decision_repo, tmp_path):
        from src.use_cases.export_papers import ExportScreenedPapersUseCase

        output = tmp_path / "output.xlsx"
        ExportScreenedPapersUseCase(mock_paper_repo, mock_decision_repo).execute(str(output))

        wb = load_workbook(output)
        ws = wb["Grey Literature"]
        comment = ws["H1"].comment
        assert comment is not None
        assert "EC1" in comment.text
        assert "EC6" in comment.text

    def test_ces2_comment_exists(self, mock_paper_repo, mock_decision_repo, tmp_path):
        from src.use_cases.export_papers import ExportScreenedPapersUseCase

        output = tmp_path / "output.xlsx"
        ExportScreenedPapersUseCase(mock_paper_repo, mock_decision_repo).execute(str(output))

        wb = load_workbook(output)
        ws = wb["Grey Literature"]
        comment = ws["K1"].comment
        assert comment is not None
        assert "EC1" in comment.text

    def test_revisor1_text_format(self, mock_paper_repo, mock_decision_repo, tmp_path):
        from src.use_cases.export_papers import ExportScreenedPapersUseCase

        output = tmp_path / "output.xlsx"
        ExportScreenedPapersUseCase(mock_paper_repo, mock_decision_repo).execute(str(output))

        wb = load_workbook(output)
        ws = wb["Grey Literature"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[8] is not None:
                assert row[8] in ("YES", "NO", "NEEDS_REVIEW")

    def test_revisor2_empty(self, mock_paper_repo, mock_decision_repo, tmp_path):
        from src.use_cases.export_papers import ExportScreenedPapersUseCase

        output = tmp_path / "output.xlsx"
        ExportScreenedPapersUseCase(mock_paper_repo, mock_decision_repo).execute(str(output))

        wb = load_workbook(output)
        ws = wb["Grey Literature"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            assert row[11] is None or row[11] == ""

    def test_decision_column_empty(self, mock_paper_repo, mock_decision_repo, tmp_path):
        from src.use_cases.export_papers import ExportScreenedPapersUseCase

        output = tmp_path / "output.xlsx"
        ExportScreenedPapersUseCase(mock_paper_repo, mock_decision_repo).execute(str(output))

        wb = load_workbook(output)
        ws = wb["Grey Literature"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            assert row[12] is None or row[12] == ""


class TestSequentialNumbering:
    def test_numbering_sequential(self, mocker: MockerFixture, tmp_path):
        from src.use_cases.export_papers import ExportScreenedPapersUseCase

        repo = mocker.MagicMock()
        repo.get_all_papers.return_value = [
            Paper(id="a", title="A", source_type=SourceType.WL, publication_year=2023),
            Paper(id="b", title="B", source_type=SourceType.GL, publication_year=2023),
            Paper(id="c", title="C", source_type=SourceType.WL, publication_year=2023),
        ]
        repo.get_all_decisions.return_value = [
            ScreeningDecision(paper_id="a", status=ScreeningStatus.INCLUDED, confidence_score=0.9, rationale="x", applied_criteria_codes=["IC1"]),
            ScreeningDecision(paper_id="b", status=ScreeningStatus.INCLUDED, confidence_score=0.9, rationale="x", applied_criteria_codes=["IC1"]),
            ScreeningDecision(paper_id="c", status=ScreeningStatus.INCLUDED, confidence_score=0.9, rationale="x", applied_criteria_codes=["IC1"]),
        ]
        repo.get_human_decision_map.return_value = {}
        repo.get_quality_map.return_value = {}

        output = tmp_path / "output.xlsx"
        ExportScreenedPapersUseCase(repo, repo).execute(str(output))

        wb = load_workbook(output)

        ws_wl = wb["White Literature"]
        wl_nums = [ws_wl.cell(row=r, column=3).value for r in range(2, ws_wl.max_row + 1)]
        assert wl_nums == [1, 2]

        ws_gl = wb["Grey Literature"]
        gl_nums = [ws_gl.cell(row=r, column=3).value for r in range(2, ws_gl.max_row + 1)]
        assert gl_nums == [1]


class TestStyling:
    def test_wl_header_fill_is_gray(self, mock_paper_repo, mock_decision_repo, tmp_path):
        from src.use_cases.export_papers import ExportScreenedPapersUseCase

        output = tmp_path / "output.xlsx"
        ExportScreenedPapersUseCase(mock_paper_repo, mock_decision_repo).execute(str(output))

        wb = load_workbook(output)
        ws = wb["White Literature"]
        cell = ws["A1"]
        assert cell.fill.start_color.rgb in ("00D9D9D9", "D9D9D9")

    def test_wl_header_bold(self, mock_paper_repo, mock_decision_repo, tmp_path):
        from src.use_cases.export_papers import ExportScreenedPapersUseCase

        output = tmp_path / "output.xlsx"
        ExportScreenedPapersUseCase(mock_paper_repo, mock_decision_repo).execute(str(output))

        wb = load_workbook(output)
        ws = wb["White Literature"]
        for cell in ws[1]:
            assert cell.font.bold is True

    def test_gl_header_bold(self, mock_paper_repo, mock_decision_repo, tmp_path):
        from src.use_cases.export_papers import ExportScreenedPapersUseCase

        output = tmp_path / "output.xlsx"
        ExportScreenedPapersUseCase(mock_paper_repo, mock_decision_repo).execute(str(output))

        wb = load_workbook(output)
        ws = wb["Grey Literature"]
        for cell in ws[1]:
            if cell.value:
                assert cell.font.bold is True

    def test_wl_title_has_wrap_text(self, mock_paper_repo, mock_decision_repo, tmp_path):
        from src.use_cases.export_papers import ExportScreenedPapersUseCase

        output = tmp_path / "output.xlsx"
        ExportScreenedPapersUseCase(mock_paper_repo, mock_decision_repo).execute(str(output))

        wb = load_workbook(output)
        ws = wb["White Literature"]
        headers = [c.value for c in ws[1]]
        title_idx = headers.index("Title") + 1
        for row in ws.iter_rows(min_row=2, min_col=title_idx, max_col=title_idx):
            for cell in row:
                assert cell.alignment.wrap_text is True

    def test_gl_title_has_wrap_text(self, mock_paper_repo, mock_decision_repo, tmp_path):
        from src.use_cases.export_papers import ExportScreenedPapersUseCase

        output = tmp_path / "output.xlsx"
        ExportScreenedPapersUseCase(mock_paper_repo, mock_decision_repo).execute(str(output))

        wb = load_workbook(output)
        ws = wb["Grey Literature"]
        for row in ws.iter_rows(min_row=2, min_col=4, max_col=4):
            for cell in row:
                assert cell.alignment.wrap_text is True


class TestEmptyDecisions:
    def test_empty_creates_both_sheets_with_only_headers(self, mock_paper_repo, mocker: MockerFixture, tmp_path):
        from src.use_cases.export_papers import ExportScreenedPapersUseCase

        decision_repo = mocker.MagicMock()
        decision_repo.get_all_decisions.return_value = []
        decision_repo.get_human_decision_map.return_value = {}
        decision_repo.get_quality_map.return_value = {}

        output = tmp_path / "output.xlsx"
        ExportScreenedPapersUseCase(mock_paper_repo, decision_repo).execute(str(output))

        wb = load_workbook(output)

        ws_wl = wb["White Literature"]
        wl_headers = [c.value for c in ws_wl[1]]
        assert len(wl_headers) > 0
        wl_data = list(ws_wl.iter_rows(min_row=2, values_only=True))
        assert all(all(c is None for c in row) for row in wl_data)

        ws_gl = wb["Grey Literature"]
        gl_headers = [c.value for c in ws_gl[1]]
        assert len(gl_headers) > 0
        gl_data = list(ws_gl.iter_rows(min_row=2, values_only=True))
        assert all(all(c is None for c in row) for row in gl_data)